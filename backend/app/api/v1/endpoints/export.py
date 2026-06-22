import csv
import io
import json
from datetime import datetime

import openpyxl
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.api.deps import get_current_active_user
from app.models.user import User
from app.models.document import Document
from app.services.document_service import get_document_by_id

router = APIRouter(prefix="/export", tags=["Export"])


def flatten_fields(extracted_fields: dict | None) -> dict:
    """
    Flatten nested extracted_fields dict into a single-level dict
    suitable for CSV/Excel rows.

    e.g. {"skills": ["React", "Node"]} → {"skills": "React, Node"}
         {"experience": [{...}]} → {"experience": "[{...}]"}
    """
    if not extracted_fields:
        return {}

    flat = {}
    for key, value in extracted_fields.items():
        if isinstance(value, list):
            if all(isinstance(v, str) for v in value):
                flat[key] = ", ".join(value)
            else:
                flat[key] = json.dumps(value)
        elif isinstance(value, dict):
            flat[key] = json.dumps(value)
        else:
            flat[key] = str(value) if value is not None else ""
    return flat


def get_export_rows(documents: list[Document]) -> tuple[list[str], list[dict]]:
    """
    Build headers and rows for export from a list of documents.
    Collects all possible field keys across all documents.
    """
    base_headers = [
        "id", "filename", "document_type", "status",
        "file_size_kb", "uploaded_at", "summary"
    ]

    # Collect all extracted field keys across all docs
    all_field_keys: set[str] = set()
    for doc in documents:
        if doc.extracted_fields:
            all_field_keys.update(doc.extracted_fields.keys())

    field_headers = sorted(all_field_keys)
    headers = base_headers + field_headers

    rows = []
    for doc in documents:
        flat_fields = flatten_fields(doc.extracted_fields)
        row = {
            "id": str(doc.id),
            "filename": doc.original_filename,
            "document_type": doc.document_type or "",
            "status": doc.status,
            "file_size_kb": round(doc.file_size / 1024, 1),
            "uploaded_at": doc.created_at.strftime("%Y-%m-%d %H:%M") if doc.created_at else "",
            "summary": doc.summary or "",
        }
        for key in field_headers:
            row[key] = flat_fields.get(key, "")
        rows.append(row)

    return headers, rows


@router.get("/csv")
def export_all_csv(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """
    Export all of the user's documents as a CSV file.
    Includes metadata + all AI-extracted fields as columns.
    """
    documents = (
        db.query(Document)
        .filter(Document.user_id == current_user.id)
        .order_by(Document.created_at.desc())
        .all()
    )

    headers, rows = get_export_rows(documents)

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    writer.writerows(rows)
    output.seek(0)

    filename = f"docintel_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/excel")
def export_all_excel(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """
    Export all of the user's documents as an Excel (.xlsx) file.
    Includes metadata + all AI-extracted fields as columns.
    """
    documents = (
        db.query(Document)
        .filter(Document.user_id == current_user.id)
        .order_by(Document.created_at.desc())
        .all()
    )

    headers, rows = get_export_rows(documents)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Documents"

    # Header row with bold styling
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    # Data rows
    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    # Auto-size columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"docintel_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/csv/{doc_id}")
def export_single_csv(
    doc_id: str,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """Export a single document's extracted fields as CSV."""
    doc = get_document_by_id(doc_id, current_user, db)
    headers, rows = get_export_rows([doc])

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    writer.writerows(rows)
    output.seek(0)

    filename = f"{doc.original_filename.rsplit('.', 1)[0]}_export.csv"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/excel/{doc_id}")
def export_single_excel(
    doc_id: str,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """Export a single document's extracted fields as Excel."""
    doc = get_document_by_id(doc_id, current_user, db)
    headers, rows = get_export_rows([doc])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = doc.original_filename[:31]

    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{doc.original_filename.rsplit('.', 1)[0]}_export.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )